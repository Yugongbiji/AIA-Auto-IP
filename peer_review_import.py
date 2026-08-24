"""Import multi-person peer reviews into the AIA Auto IP profile layer."""
from __future__ import annotations
import argparse, json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
import openpyxl
import server

EXPECTED_COLUMNS = {
    "nickname": "1. 你平时怎么称呼TA？", "relationship": "2.你和TA的关系？",
    "traits": "3. 如果只能选3个词形容TA，你会选？", "topics": "4. 遇到什么事情，你比较愿意找TA聊聊？",
    "roles": "5. 如果把TA介绍给朋友，你觉得TA更像哪种人？", "intro": "6. 如果只能用一句话把TA介绍给一个不认识的人，你会怎么说？",
    "agentId": "7.请输入分享给你问卷的营销员信息",
}
def clean(value) -> str: return str(value or "").strip()
def split_multi(value) -> list[str]:
    text = clean(value).replace("；", ";").replace("、", ";").replace("，", ";").replace(",", ";")
    return [item.strip() for item in text.split(";") if item.strip() and item.strip() != "其他"]
def count_items(counter: Counter) -> list[dict]:
    """Keep every distinct raw answer with its frequency; UI controls visual collapsing."""
    return [{"label": label, "count": count} for label, count in counter.most_common()]

def aggregate(rows: list[dict]) -> dict:
    nicknames, relationships, traits, topics, roles = Counter(), Counter(), Counter(), Counter(), Counter(); quotes = []
    for row in rows:
        if row["nickname"]: nicknames[row["nickname"]] += 1
        if row["relationship"]: relationships[row["relationship"]] += 1
        traits.update(split_multi(row["traits"])); topics.update(split_multi(row["topics"])); roles.update(split_multi(row["roles"]))
        intro = row["intro"]
        if intro and intro not in quotes: quotes.append(intro)
    return {"source":"身边人评价问卷","reviewCount":len(rows),"topNicknames":count_items(nicknames),"relationships":count_items(relationships),"topTraits":count_items(traits),"topTopics":count_items(topics),"topRoles":count_items(roles),"representativeQuotes":quotes}

def summary_text(summary: dict) -> str:
    nicknames = "、".join(f'{item["label"]}×{item["count"]}' for item in summary.get("topNicknames", [])[:4])
    traits = "、".join(item["label"] for item in summary.get("topTraits", [])[:5]); roles = "、".join(item["label"] for item in summary.get("topRoles", [])[:4]); topics = "、".join(item["label"] for item in summary.get("topTopics", [])[:4]); parts=[]
    if nicknames: parts.append(f"常用称呼：{nicknames}")
    if traits: parts.append(f"高频印象：{traits}")
    if roles: parts.append(f"他人角色认知：{roles}")
    if topics: parts.append(f"愿意咨询的话题：{topics}")
    return "；".join(parts)

def read_workbook(path: Path) -> dict[str,list[dict]]:
    workbook=openpyxl.load_workbook(path,read_only=True,data_only=True); target=None
    for sheet in workbook.worksheets:
        headers=[clean(cell.value) for cell in next(sheet.iter_rows(min_row=1,max_row=1))]
        if EXPECTED_COLUMNS["agentId"] in headers: target=sheet; break
    if target is None: raise RuntimeError("没有找到包含第 7 题营销员编号的评价明细工作表。")
    rows=target.iter_rows(values_only=True); headers=[clean(value) for value in next(rows)]; index={header:position for position,header in enumerate(headers)}
    missing=[header for header in EXPECTED_COLUMNS.values() if header not in index]
    if missing: raise RuntimeError("评价表缺少固定列："+"；".join(missing))
    grouped=defaultdict(list)
    for raw in rows:
        record={key:clean(raw[index[header]]) if index[header]<len(raw) else "" for key,header in EXPECTED_COLUMNS.items()}; agent_id=record["agentId"]
        if agent_id: grouped[agent_id].append(record)
    return grouped

def ensure_peer_review_table(conn):
    id_column="INTEGER PRIMARY KEY AUTOINCREMENT" if server.database_engine()=="sqlite" else "BIGSERIAL PRIMARY KEY"
    conn.executescript(f"""CREATE TABLE IF NOT EXISTS peer_reviews (id {id_column},agent_id TEXT NOT NULL,reviewer_nickname TEXT,relationship TEXT,traits TEXT,topics TEXT,roles TEXT,intro TEXT,imported_at TEXT NOT NULL,FOREIGN KEY(agent_id) REFERENCES agents(agent_id)); CREATE INDEX IF NOT EXISTS idx_peer_reviews_agent_id ON peer_reviews(agent_id,id);""")
def merge_into_saved_profile(conn,agent_id:str,summary:dict,now:str):
    row=conn.execute("SELECT profile_json FROM saved_profiles WHERE agent_id = ?",(agent_id,)).fetchone(); profile_json=row["profile_json"] if row else ""; profile=json.loads(profile_json) if profile_json else {}
    profile["peerReviewSummary"]=summary; profile["peerReviewKeywords"]=summary_text(summary)
    conn.execute("""INSERT INTO saved_profiles(agent_id,profile_json,updated_at) VALUES (?,?,?) ON CONFLICT(agent_id) DO UPDATE SET profile_json=excluded.profile_json,updated_at=excluded.updated_at""",(agent_id,json.dumps(profile,ensure_ascii=False),now))
def import_peer_reviews(path:Path)->dict:
    grouped=read_workbook(path); now=datetime.now(timezone.utc).isoformat(); imported_reviews=0; matched_agents=0; skipped_agents=[]
    with server.database() as conn:
        ensure_peer_review_table(conn)
        for agent_id,rows in grouped.items():
            if not conn.execute("SELECT 1 FROM agents WHERE agent_id = ?",(agent_id,)).fetchone(): skipped_agents.append(agent_id); continue
            conn.execute("DELETE FROM peer_reviews WHERE agent_id = ?",(agent_id,))
            for row in rows:
                conn.execute("INSERT INTO peer_reviews(agent_id,reviewer_nickname,relationship,traits,topics,roles,intro,imported_at) VALUES (?,?,?,?,?,?,?,?)",(agent_id,row["nickname"],row["relationship"],row["traits"],row["topics"],row["roles"],row["intro"],now)); imported_reviews+=1
            merge_into_saved_profile(conn,agent_id,aggregate(rows),now); matched_agents+=1
    return {"matchedAgents":matched_agents,"importedReviews":imported_reviews,"skippedAgentIds":skipped_agents}
def main():
    parser=argparse.ArgumentParser(); parser.add_argument("xlsx",type=Path); args=parser.parse_args(); server.load_local_env(); server.initialize_database(); print(json.dumps(import_peer_reviews(args.xlsx),ensure_ascii=False,indent=2))
if __name__=="__main__": main()
