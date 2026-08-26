#!/usr/bin/env python3
"""Import the human-approved 57-person IP baseline from an Excel review file.

Default mode is dry-run. Use --commit only after matched/missing/invalid counts are clean.
The workbook itself is never committed to GitHub.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import server
from backend.stable_ip import ensure_stable_schema, promote_output, validate_output

SHEET_CANDIDATES = ("57人推荐简介", "57人最终合规验收")
ALIASES = {
    "agent_id": ("营销员编号",),
    "nickname": ("人工确认昵称",),
    "who": ("我是谁（事实层）", "我是谁（原始事实层）", "我是谁"),
    "advantage": ("我的优势（事实层）", "我的优势（原始事实层）", "我的优势"),
    "value": ("我能提供什么价值（151重做）", "我能提供什么价值（原始事实层）", "我能提供什么价值"),
    "headline": ("Canonical一句话IP", "合规后Canonical一句话IP", "一句话IP"),
    "xhs": ("小红书推荐简介（合规完整版）", "小红书推荐简介（完整）"),
    "video": ("视频号/抖音推荐简介（完整版）", "视频号/抖音人物正文+IP"),
}


def clean(value) -> str:
    return str(value or "").strip()


def lines(value) -> list[str]:
    return [x.strip() for x in clean(value).splitlines() if x.strip()]


def choose_sheet(book):
    for name in SHEET_CANDIDATES:
        if name in book.sheetnames:
            return book[name]
    return book[book.sheetnames[0]]


def header_map(sheet) -> dict[str, int]:
    raw = {clean(cell.value): idx for idx, cell in enumerate(sheet[1]) if clean(cell.value)}
    resolved = {}
    for key, aliases in ALIASES.items():
        for alias in aliases:
            if alias in raw:
                resolved[key] = raw[alias]
                break
    missing = [key for key in ALIASES if key not in resolved]
    if missing:
        raise RuntimeError("Excel 缺少必要列：" + ", ".join(missing))
    return resolved


def build_output(row, cols):
    def v(key):
        return row[cols[key]].value if cols[key] < len(row) else ""
    return {
        "nickname": clean(v("nickname")),
        "whoLines": lines(v("who")),
        "advantageLines": lines(v("advantage")),
        "valueLines": lines(v("value")),
        "headline": clean(v("headline")),
        "xiaohongshuBio": lines(v("xhs")),
        "videoDouyinBio": lines(v("video")),
    }


def load_rows(path: Path):
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = choose_sheet(book)
    cols = header_map(sheet)
    records = []
    for row in sheet.iter_rows(min_row=2):
        agent_id = clean(row[cols["agent_id"]].value)
        if not agent_id:
            continue
        records.append((agent_id, build_output(row, cols)))
    return records


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--commit", action="store_true", help="通过校验后正式写入数据库")
    parser.add_argument("--expected", type=int, default=57)
    parser.add_argument("--source", default="human_approved_baseline_20260826")
    args = parser.parse_args()

    server.load_local_env()
    server.initialize_database()
    records = load_rows(args.xlsx)
    duplicates = sorted({aid for aid, _ in records if sum(1 for x, _ in records if x == aid) > 1})
    invalid = [(aid, validate_output(output)) for aid, output in records if validate_output(output)]

    with server.database() as conn:
        ensure_stable_schema(conn)
        known = {row["agent_id"] for row in conn.execute("SELECT agent_id FROM agents").fetchall()}
        missing = sorted(aid for aid, _ in records if aid not in known)

    summary = {
        "rows": len(records),
        "expected": args.expected,
        "matched": len(records) - len(missing),
        "missing": missing,
        "duplicates": duplicates,
        "invalid": invalid,
        "mode": "commit" if args.commit else "dry-run",
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if len(records) != args.expected or missing or duplicates or invalid:
        raise SystemExit("校验未通过，未写入数据库。")
    if not args.commit:
        print("DRY-RUN 通过。确认后加 --commit 正式导入。")
        return

    promoted = []
    with server.database() as conn:
        ensure_stable_schema(conn)
        for agent_id, output in records:
            result = promote_output(conn, agent_id, output, args.source, approved=True)
            if not result.get("promoted"):
                raise RuntimeError(f"{agent_id} 导入失败：{result}")
            promoted.append({"agentId": agent_id, **result})
    print(json.dumps({"committed": len(promoted), "records": promoted}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
