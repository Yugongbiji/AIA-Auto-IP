#!/usr/bin/env python3
"""AIA Auto IP 脚本库 Excel 导入器 V1。

默认只做 dry-run；显式传 --write 才写数据库。
兼容 PostgreSQL（DATABASE_URL）与 Preview SQLite。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SPEECH_RATE = 260


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def clean_tag(value):
    return re.sub(r"\s+", "", clean_text(value))


def parse_titles(raw):
    text = clean_text(raw)
    if not text:
        return ["", "", ""]
    # 先识别 1/2/3 编号；编号可能同行，也可能换行。
    pattern = re.compile(r"(?:^|\s|\n)([123])\s*[、\.．:：]\s*")
    matches = list(pattern.finditer(text))
    titles = []
    if matches:
        for i, match in enumerate(matches[:3]):
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            title = text[match.end():end].strip(" \n\t;；")
            if title:
                titles.append(title)
    else:
        titles = [x.strip() for x in text.split("\n") if x.strip()]
    if not titles:
        titles = [text]
    return (titles + ["", "", ""])[:3]


def count_spoken_chars(body):
    # 统计中英文/数字口播字符，不把标点、空白计入口播字数。
    return len(re.findall(r"[\u4e00-\u9fffA-Za-z0-9]", body))


def content_hash(body):
    normalized = re.sub(r"\s+", "", body)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value) or None


def load_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [clean_text(c.value) for c in next(ws.iter_rows())]
    index = {name: i for i, name in enumerate(headers)}
    required = ["一级标签", "二级标签", "标题", "正文"]
    missing = [x for x in required if x not in index]
    if missing:
        raise ValueError(f"缺少必要列: {', '.join(missing)}")

    result = []
    for row_no, row in enumerate(ws.iter_rows(), start=2):
        vals = [c.value for c in row]
        body = clean_text(vals[index["正文"]])
        if not body:
            continue
        titles = parse_titles(vals[index["标题"]])
        wc = count_spoken_chars(body)
        result.append({
            "source_row": row_no,
            "batch": clean_text(vals[index["批次"]]) if "批次" in index else "",
            "level1_tag": clean_tag(vals[index["一级标签"]]),
            "level2_tag": clean_tag(vals[index["二级标签"]]),
            "title_1": titles[0], "title_2": titles[1] or None, "title_3": titles[2] or None,
            "body": body,
            "word_count": wc,
            "estimated_minutes": round(wc / SPEECH_RATE, 1),
            "is_hot": clean_tag(vals[index["一级标签"]]) == "热点",
            "reviewed_at": normalize_date(vals[index["过审时间"]]) if "过审时间" in index else None,
            "content_hash": content_hash(body),
        })
    return result


def audit(rows):
    seen = {}
    duplicate_rows = []
    for item in rows:
        if item["content_hash"] in seen:
            duplicate_rows.append((item["source_row"], seen[item["content_hash"]]))
        else:
            seen[item["content_hash"]] = item["source_row"]
    return {
        "source_rows": len(rows),
        "unique_scripts": len(seen),
        "duplicate_rows": duplicate_rows,
        "missing_title_1": sum(not r["title_1"] for r in rows),
        "missing_title_2": sum(not r["title_2"] for r in rows),
        "missing_title_3": sum(not r["title_3"] for r in rows),
    }


def schema_sql(dialect):
    pk = "BIGSERIAL PRIMARY KEY" if dialect == "postgres" else "INTEGER PRIMARY KEY AUTOINCREMENT"
    bool_type = "BOOLEAN" if dialect == "postgres" else "INTEGER"
    return f"""
CREATE TABLE IF NOT EXISTS script_library (
  script_id {pk}, batch TEXT, level1_tag TEXT NOT NULL, level2_tag TEXT,
  title_1 TEXT NOT NULL, title_2 TEXT, title_3 TEXT, body TEXT NOT NULL,
  word_count INTEGER NOT NULL, estimated_minutes REAL NOT NULL,
  is_hot {bool_type} NOT NULL DEFAULT 0, reviewed_at TEXT,
  status TEXT NOT NULL DEFAULT 'active', content_hash TEXT NOT NULL UNIQUE,
  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_script_library_tags ON script_library(level1_tag, level2_tag);
CREATE INDEX IF NOT EXISTS idx_script_library_status ON script_library(status);
"""


def write_sqlite(rows, db_path):
    conn = sqlite3.connect(db_path)
    conn.executescript(schema_sql("sqlite"))
    sql = """INSERT OR IGNORE INTO script_library
(batch,level1_tag,level2_tag,title_1,title_2,title_3,body,word_count,estimated_minutes,is_hot,reviewed_at,content_hash)
VALUES (?,?,?,?,?,?,?,?,?,?,?,?)"""
    for r in rows:
        conn.execute(sql, (r["batch"],r["level1_tag"],r["level2_tag"],r["title_1"],r["title_2"],r["title_3"],r["body"],r["word_count"],r["estimated_minutes"],int(r["is_hot"]),r["reviewed_at"],r["content_hash"]))
    conn.commit(); conn.close()


def main():
    p = argparse.ArgumentParser()
    p.add_argument("excel", type=Path)
    p.add_argument("--write", action="store_true", help="确认后才写入数据库")
    p.add_argument("--sqlite", type=Path, help="写入 Preview SQLite；V1 默认优先用于 Preview 验证")
    p.add_argument("--report", type=Path)
    args = p.parse_args()
    rows = load_rows(args.excel)
    report = audit(rows)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if args.report:
        args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.write:
        if not args.sqlite:
            raise SystemExit("V1 安全限制：--write 必须显式提供 --sqlite；正式 RDS 写入待 Preview 验收后开放。")
        write_sqlite(rows, args.sqlite)
        print(f"已写入 Preview SQLite: {args.sqlite}")
    else:
        print("dry-run 完成：未写入任何数据库。")

if __name__ == "__main__":
    main()
