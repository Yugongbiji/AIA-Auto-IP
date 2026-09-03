#!/usr/bin/env python3
"""Safe one-time importer for Nanjing Huijia 2026-09-03 batch.

Inputs stay outside Git. Dry-run is default. --commit creates a SQLite backup first.
Existing agents are never overwritten by this batch; existing current_ip_outputs are a hard conflict.
Workbook discovery is content-based so Linux ZIP filename mojibake cannot mis-route files.
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import server
from backend.stable_ip import ensure_stable_schema, quality_score, validate_output

XHS_FOOTER = "本账号所述内容为个人意见，不代表任何官方意见。"
SOURCE = "nanjing_huijia_human_approved_20260903"


def clean(v): return str(v or "").strip()
def split_lines(v): return [x.strip() for x in clean(v).splitlines() if x.strip()]


def workbook_signature(path: Path) -> set[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sig = set(wb.sheetnames)
    for ws in wb.worksheets[:2]:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            sig.update(clean(x) for x in row if clean(x))
    wb.close()
    return sig


def discover_workbooks(folder: Path) -> dict[str, Path]:
    found = {}
    details = {}
    for path in folder.rglob("*.xlsx"):
        try:
            sig = workbook_signature(path)
        except Exception as exc:
            details[path.name] = f"unreadable:{exc}"
            continue
        joined = "\n".join(sig)
        kind = None
        if "稳定输出" in sig and "推荐昵称" in sig and "一句话IP" in sig:
            kind = "stable"
        elif "最终名单" in sig or ("直辖组" in sig and "姓名" in sig and "序号" in sig and "工号" in sig):
            kind = "roster"
        elif "7.请输入分享给你问卷的营销员信息" in sig and ("2.你和TA的关系？" in sig or "3. 如果只能选3个词形容TA，你会选？" in sig):
            kind = "reviews"
        elif "营销员编号" in sig and "姓名" in sig and ("主要目的" in joined or "自媒体" in joined):
            kind = "basic"
        details[path.name] = kind or "unknown"
        if kind:
            if kind in found:
                raise RuntimeError(f"识别到多个 {kind} 文件：{found[kind].name}；{path.name}")
            found[kind] = path
    missing = [k for k in ("roster", "basic", "reviews", "stable") if k not in found]
    if missing:
        raise RuntimeError("无法按内容识别全部4类Excel；missing=" + ",".join(missing) + "；details=" + json.dumps(details, ensure_ascii=False))
    return found


def load_roster(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb["最终名单"] if "最终名单" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    for row in rows:
        vals = [clean(x) for x in row]
        if "序号" in vals and ("工号" in vals or "营销员编号" in vals):
            header = vals; break
    else: raise RuntimeError("最终名单未找到表头")
    idx = {h:i for i,h in enumerate(header)}; aid_key = "工号" if "工号" in idx else "营销员编号"
    out={}
    for row in rows:
        aid=clean(row[idx[aid_key]]) if idx[aid_key] < len(row) else ""; name=clean(row[idx["姓名"]]) if idx["姓名"] < len(row) else ""; group=clean(row[idx["直辖组"]]) if "直辖组" in idx and idx["直辖组"] < len(row) else ""
        if aid and name: out[aid]={"agentId":aid,"name":name,"directGroup":group}
    return out


def load_basic(path: Path, roster: dict):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb.worksheets[0]; rows=ws.iter_rows(values_only=True); headers=[clean(x) for x in next(rows)]; keys=[server.header_key(h) for h in headers]; out={}
    for row in rows:
        rec={k:clean(v) for k,v in zip(keys,row) if k and clean(v)}; aid=rec.get("agentId","")
        if aid in roster and rec.get("name")==roster[aid]["name"]: rec["directGroup"]=roster[aid]["directGroup"]; out[aid]=rec
    return out


def load_reviews(path: Path, roster: dict):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb.worksheets[0]; rows=ws.iter_rows(values_only=True); headers=[clean(x) for x in next(rows)]; idx={h:i for i,h in enumerate(headers)}
    aliases={"nickname":["1. 生活中你怎么称呼TA","1. 你平时怎么称呼TA？"],"relationship":["2.你和TA的关系？"],"traits":["3. 如果只能选3个词形容TA，你会选？"],"topics":["4. 遇到什么事情，你比较愿意找TA聊聊？"],"roles":["5. 如果把TA介绍给朋友，你觉得TA更像哪种人？"],"intro":["6. 如果只能用一句话把TA介绍给一个不认识的人，你会怎么说？"],"agentId":["7.请输入分享给你问卷的营销员信息"]}
    cols={}
    for key,names in aliases.items():
        cols[key]=next((idx[n] for n in names if n in idx),None)
        if cols[key] is None: raise RuntimeError(f"客户反馈表缺列：{key}")
    grouped=defaultdict(list)
    for row in rows:
        aid=clean(row[cols["agentId"]]);
        if aid in roster: grouped[aid].append({k:clean(row[pos]) for k,pos in cols.items() if k!="agentId"})
    return grouped


def split_multi(v):
    text=clean(v).replace("；",";").replace("、",";").replace("，",";").replace(",",";"); return [x.strip() for x in text.split(";") if x.strip() and x.strip()!="其他"]

def review_summary(rows):
    nick,rel,traits,topics,roles=Counter(),Counter(),Counter(),Counter(),Counter(); quotes=[]
    for r in rows:
        if r["nickname"]: nick[r["nickname"]]+=1
        if r["relationship"]: rel[r["relationship"]]+=1
        traits.update(split_multi(r["traits"])); topics.update(split_multi(r["topics"])); roles.update(split_multi(r["roles"]));
        if r["intro"] and r["intro"] not in quotes: quotes.append(r["intro"])
    fmt=lambda c:[{"label":k,"count":v} for k,v in c.most_common()]
    return {"source":"身边人评价问卷","reviewCount":len(rows),"topNicknames":fmt(nick),"relationships":fmt(rel),"topTraits":fmt(traits),"topTopics":fmt(topics),"topRoles":fmt(roles),"representativeQuotes":quotes}


def load_stable(path: Path, roster: dict):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb["稳定输出"]; rows=ws.iter_rows(values_only=True); headers=[clean(x) for x in next(rows)]; idx={h:i for i,h in enumerate(headers)}
    need=["工号","姓名","生成状态","推荐昵称","一句话IP","小红书推荐简介（含合规声明）","视频号/抖音推荐简介（含合规声明）"]
    missing=[h for h in need if h not in idx]
    if missing: raise RuntimeError("稳定稿缺列："+"；".join(missing))
    out={}
    for row in rows:
        aid=clean(row[idx["工号"]]); status=clean(row[idx["生成状态"]])
        if aid not in roster or not status.startswith("已生成"): continue
        if clean(row[idx["姓名"]])!=roster[aid]["name"]: raise RuntimeError(f"稳定稿姓名不匹配：{aid}")
        xhs=split_lines(row[idx["小红书推荐简介（含合规声明）"]]); video=split_lines(row[idx["视频号/抖音推荐简介（含合规声明）"]]); headline=clean(row[idx["一句话IP"]]); body=[x for x in xhs if x!=headline and x!=XHS_FOOTER]
        advantages=[x for x in body if x.startswith(("✨","🌟","🏆","⭐"))]; who=[x for x in body if x not in advantages]
        if not who and body: advantages=body
        out[aid]={"nickname":clean(row[idx["推荐昵称"]]),"whoLines":who,"advantageLines":advantages,"valueLines":[],"headline":headline,"xiaohongshuBio":xhs,"videoDouyinBio":video}
    return out


def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--input-dir",type=Path); ap.add_argument("--roster",type=Path); ap.add_argument("--basic",type=Path); ap.add_argument("--reviews",type=Path); ap.add_argument("--stable",type=Path); ap.add_argument("--commit",action="store_true"); args=ap.parse_args()
    if args.input_dir:
        files=discover_workbooks(args.input_dir); args.roster=files["roster"]; args.basic=files["basic"]; args.reviews=files["reviews"]; args.stable=files["stable"]
        print(json.dumps({"detectedFiles":{k:v.name for k,v in files.items()}},ensure_ascii=False,indent=2))
    if not all((args.roster,args.basic,args.reviews,args.stable)): raise SystemExit("请提供 --input-dir，或完整提供4个Excel参数。")
    server.load_local_env(); server.initialize_database()
    if server.database_engine()!="sqlite": raise SystemExit("本批次脚本只允许在当前 SQLite Production/Preview 数据库执行。")
    roster=load_roster(args.roster); basic=load_basic(args.basic,roster); reviews=load_reviews(args.reviews,roster); stable=load_stable(args.stable,roster); invalid={aid:validate_output(v) for aid,v in stable.items() if validate_output(v)}
    with server.database() as conn:
        ensure_stable_schema(conn); existing_agents={r["agent_id"] for r in conn.execute("SELECT agent_id FROM agents").fetchall()}; stable_conflicts={r["agent_id"] for r in conn.execute("SELECT agent_id FROM current_ip_outputs").fetchall()} & set(roster); before={t:conn.execute(f"SELECT COUNT(*) n FROM {t}").fetchone()["n"] for t in ("agents","saved_profiles","peer_reviews","proposals","current_ip_outputs")}
    summary={"mode":"commit" if args.commit else "dry-run","roster":len(roster),"basic":len(basic),"reviewAgents":len(reviews),"reviewRows":sum(map(len,reviews.values())),"stable":len(stable),"newAgents":len(set(roster)-existing_agents),"existingRosterAgents":sorted(set(roster)&existing_agents),"stableConflicts":sorted(stable_conflicts),"invalidStable":invalid,"before":before}
    print(json.dumps(summary,ensure_ascii=False,indent=2))
    if len(roster)!=57 or len(stable)!=41 or stable_conflicts or invalid: raise SystemExit("DRY-RUN 门禁未通过，禁止写库。")
    if not args.commit: print("DRY-RUN 通过；加 --commit 才会写库。"); return
    ts=datetime.now(timezone.utc).isoformat(); backup=server.DB_PATH.with_name(f"persona.sqlite3.bak-nanjing-{datetime.now().strftime('%Y%m%d-%H%M%S')}"); shutil.copy2(server.DB_PATH,backup)
    try:
        with server.database() as conn:
            for aid,item in roster.items():
                if conn.execute("SELECT 1 FROM agents WHERE agent_id=?",(aid,)).fetchone(): continue
                survey=basic.get(aid) or {"name":item["name"],"agentId":aid,"directGroup":item["directGroup"]}; conn.execute("INSERT INTO agents(agent_id,name,survey_json,imported_at) VALUES (?,?,?,?)",(aid,item["name"],json.dumps(survey,ensure_ascii=False),ts))
            for aid,rows in reviews.items():
                if conn.execute("SELECT COUNT(*) n FROM peer_reviews WHERE agent_id=?",(aid,)).fetchone()["n"]: raise RuntimeError(f"{aid} 已存在 peer_reviews，停止以避免覆盖")
                for r in rows: conn.execute("INSERT INTO peer_reviews(agent_id,reviewer_nickname,relationship,traits,topics,roles,intro,imported_at) VALUES (?,?,?,?,?,?,?,?)",(aid,r["nickname"],r["relationship"],r["traits"],r["topics"],r["roles"],r["intro"],ts))
            for aid in set(basic)|set(reviews):
                if conn.execute("SELECT 1 FROM saved_profiles WHERE agent_id=?",(aid,)).fetchone(): raise RuntimeError(f"{aid} 已存在 saved_profiles，停止以避免覆盖")
                profile=dict(basic.get(aid) or {});
                if aid in reviews: profile["peerReviewSummary"]=review_summary(reviews[aid])
                conn.execute("INSERT INTO saved_profiles(agent_id,profile_json,updated_at) VALUES (?,?,?)",(aid,json.dumps(profile,ensure_ascii=False),ts))
            for aid,output in stable.items():
                if conn.execute("SELECT 1 FROM current_ip_outputs WHERE agent_id=?",(aid,)).fetchone(): raise RuntimeError(f"{aid} 已存在稳定稿，停止以避免覆盖")
                version=conn.execute("SELECT COALESCE(MAX(version),0)+1 n FROM proposals WHERE agent_id=?",(aid,)).fetchone()["n"]; score=quality_score(output); payload=dict(output); payload["_stableMeta"]={"source":SOURCE,"qualityScore":score,"approved":True}; conn.execute("INSERT INTO proposals(agent_id,version,proposal_json,model,created_at) VALUES (?,?,?,?,?)",(aid,version,json.dumps(payload,ensure_ascii=False),"human-approved",ts)); conn.execute("INSERT INTO current_ip_outputs(agent_id,proposal_version,output_json,quality_score,source,approved_at,updated_at) VALUES (?,?,?,?,?,?,?)",(aid,version,json.dumps(output,ensure_ascii=False),score,SOURCE,ts,ts))
        with server.database() as conn:
            after={t:conn.execute(f"SELECT COUNT(*) n FROM {t}").fetchone()["n"] for t in ("agents","saved_profiles","peer_reviews","proposals","current_ip_outputs")}; batch_stable=conn.execute("SELECT COUNT(*) n FROM current_ip_outputs WHERE source=?",(SOURCE,)).fetchone()["n"]
        print(json.dumps({"committed":True,"backup":str(backup),"after":after,"batchStable":batch_stable},ensure_ascii=False,indent=2))
        if batch_stable!=41: raise RuntimeError("写库后南京稳定稿数量不是41")
    except Exception:
        shutil.copy2(backup,server.DB_PATH); print(f"IMPORT FAILED; database restored from {backup}",file=sys.stderr); raise

if __name__=="__main__": main()
